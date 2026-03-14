#!/usr/bin/env python3
"""
BLS OEWS Data Integration Script (Python version - faster for large Excel files)

Reads all_data_M_2024.xlsx and maps BLS salary data to Pulsafi jobs and cities.

Prerequisites:
  pip install openpyxl

Usage: python3 scripts/fetch-bls-data.py
"""

import json
import os
import sys
from pathlib import Path

# Try to import openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import load_workbook

# ─── SOC Code to Job Slug mapping ───
SOC_TO_SLUG = {
    "15-1252": "software-engineer",
    "15-2051": "data-scientist",
    "15-1212": "cybersecurity-analyst",
    "15-1244": "devops-engineer",
    "15-1255": "web-developer",
    "15-1242": "database-administrator",
    "11-3021": "it-manager",
    "15-1211": "systems-analyst",
    "15-1241": "network-engineer",
    "15-1253": "qa-engineer",
    "27-3042": "technical-writer",
    "15-1256": "mobile-developer",
    "15-1243": "data-engineer",
    "15-1232": "it-support-specialist",
    "15-1299": "solutions-architect",
    "15-1221": "scrum-master",
    "15-1254": "frontend-developer",
    "15-1251": "backend-developer",
    "29-1141": "registered-nurse",
    "29-1228": "physician",
    "29-1021": "dentist",
    "29-1051": "pharmacist",
    "29-1123": "physical-therapist",
    "29-1171": "nurse-practitioner",
    "31-9092": "medical-assistant",
    "29-1292": "dental-hygienist",
    "29-2034": "radiologic-technologist",
    "29-1122": "occupational-therapist",
    "29-1127": "speech-pathologist",
    "29-1248": "surgeon",
    "29-1211": "anesthesiologist",
    "29-1223": "psychiatrist",
    "29-1041": "optometrist",
    "29-1131": "veterinarian",
    "29-2042": "emt-paramedic",
    "29-2012": "medical-lab-technician",
    "29-1126": "respiratory-therapist",
    "29-1031": "dietitian",
    "29-9021": "health-information-technician",
    "31-1120": "home-health-aide",
    "29-2055": "surgical-technologist",
    "31-9097": "phlebotomist",
    "29-2072": "medical-records-specialist",
    "13-2011": "accountant",
    "13-2051": "financial-analyst",
    "11-2021": "marketing-manager",
    "11-3121": "human-resources-manager",
    "13-1111": "management-consultant",
    "11-9199": "project-manager",
    "11-1021": "operations-manager",
    "11-2022": "sales-manager",
    "13-2052": "financial-advisor",
    "15-2011": "actuary",
    "19-3011": "economist",
    "13-2031": "budget-analyst",
    "13-2072": "loan-officer",
    "13-2053": "insurance-underwriter",
    "41-9022": "real-estate-agent",
    "43-4131": "personal-banker",
    "13-2082": "tax-preparer",
    "13-1041": "compliance-officer",
    "13-2054": "risk-analyst",
    "11-3061": "purchasing-manager",
    "13-1161": "market-research-analyst",
    "11-1011": "chief-financial-officer",
    "25-2021": "elementary-teacher",
    "25-2031": "high-school-teacher",
    "25-2059": "special-education-teacher",
    "21-1012": "school-counselor",
    "11-9032": "principal",
    "25-1099": "college-professor",
    "25-9045": "teaching-assistant",
    "25-4022": "librarian",
    "25-9031": "instructional-designer",
    "11-9039": "education-administrator",
    "25-3011": "esl-teacher",
    "25-2011": "preschool-teacher",
    "19-3034": "school-psychologist",
    "25-9099": "tutoring-coordinator",
    "17-2141": "mechanical-engineer",
    "17-2071": "electrical-engineer",
    "17-2051": "civil-engineer",
    "17-2041": "chemical-engineer",
    "17-2011": "aerospace-engineer",
    "17-2112": "industrial-engineer",
    "17-2081": "environmental-engineer",
    "17-2031": "biomedical-engineer",
    "17-2171": "petroleum-engineer",
    "17-2131": "materials-engineer",
    "17-2161": "nuclear-engineer",
    "17-2021": "agricultural-engineer",
    "17-2121": "marine-engineer",
    "17-2151": "mining-engineer",
    "17-2199": "robotics-engineer",
    "17-2111": "fire-protection-engineer",
    "47-2111": "electrician",
    "47-2152": "plumber",
    "49-9021": "hvac-technician",
    "47-2031": "carpenter",
    "51-4121": "welder",
    "47-2073": "heavy-equipment-operator",
    "11-9021": "construction-manager",
    "49-3023": "auto-mechanic",
    "49-3031": "diesel-mechanic",
    "47-4021": "elevator-installer",
    "47-2171": "ironworker",
    "47-2021": "mason",
    "47-2141": "painter",
    "47-2181": "roofer",
    "47-2211": "sheet-metal-worker",
    "23-1011": "lawyer",
    "23-2011": "paralegal",
    "43-6012": "legal-secretary",
    "23-1023": "judge",
    "23-1022": "mediator",
    "23-2099": "legal-analyst",
    "27-1024": "graphic-designer",
    "27-4032": "video-editor",
    "27-3043": "content-writer",
    "27-4021": "photographer",
    "27-1011": "art-director",
    "27-1014": "animator",
    "27-3031": "public-relations-specialist",
    "27-3023": "journalist",
    "27-2012": "film-director",
    "27-2041": "music-producer",
    "33-3051": "police-officer",
    "33-2011": "firefighter",
    "21-1021": "social-worker",
    "19-3051": "urban-planner",
    "33-3021": "federal-agent",
    "19-2041": "environmental-scientist",
    "21-1091": "public-health-advisor",
}

# Extra mappings for jobs sharing SOC codes
EXTRA_MAPPINGS = {
    "full-stack-developer": "15-1252",
    "game-developer": "15-1255",
    "blockchain-developer": "15-1299",
    "ai-engineer": "15-2051",
    "business-analyst": "13-1111",
    "auditor": "13-2011",
    "investment-banker": "13-2051",
    "compliance-specialist": "13-1041",
    "corporate-counsel": "23-1011",
    "patent-attorney": "23-1011",
    "public-defender": "23-1011",
    "social-media-manager": "27-3031",
    "copywriter": "27-3043",
    "structural-engineer": "17-2051",
    "automotive-engineer": "17-2141",
    "manufacturing-engineer": "17-2112",
    "geological-engineer": "17-2151",
    "curriculum-developer": "25-9031",
    "intelligence-analyst": "33-3021",
    "customs-officer": "33-3051",
    "city-manager": "11-1011",
}

# Build reverse: slug -> SOC
SLUG_TO_SOC = {}
for soc, slug in SOC_TO_SLUG.items():
    if slug not in SLUG_TO_SOC:
        SLUG_TO_SOC[slug] = soc
for slug, soc in EXTRA_MAPPINGS.items():
    if slug not in SLUG_TO_SOC:
        SLUG_TO_SOC[slug] = soc

# ─── Metro CBSA to city slugs ───
METRO_TO_CITIES = {
    "35620": ["new-york-ny", "newark-nj", "jersey-city-nj"],
    "31080": ["los-angeles-ca", "long-beach-ca", "anaheim-ca", "glendale-ca", "pasadena-ca", "torrance-ca", "inglewood-ca", "burbank-ca"],
    "16980": ["chicago-il", "naperville-il", "evanston-il"],
    "19100": ["dallas-tx", "fort-worth-tx", "arlington-tx", "plano-tx", "irving-tx", "garland-tx", "frisco-tx", "mckinney-tx", "denton-tx"],
    "26420": ["houston-tx", "sugar-land-tx", "pasadena-tx", "pearland-tx", "league-city-tx"],
    "47900": ["washington-dc", "arlington-va", "alexandria-va"],
    "33100": ["miami-fl", "fort-lauderdale-fl", "hialeah-fl", "hollywood-fl", "coral-springs-fl", "pembroke-pines-fl", "boca-raton-fl"],
    "37980": ["philadelphia-pa", "camden-nj"],
    "12060": ["atlanta-ga", "sandy-springs-ga", "roswell-ga", "alpharetta-ga", "marietta-ga"],
    "14460": ["boston-ma", "cambridge-ma"],
    "38060": ["phoenix-az", "mesa-az", "chandler-az", "scottsdale-az", "tempe-az", "gilbert-az", "glendale-az", "peoria-az"],
    "40140": ["riverside-ca", "san-bernardino-ca", "ontario-ca", "fontana-ca", "moreno-valley-ca", "corona-ca"],
    "19740": ["denver-co", "aurora-co", "lakewood-co"],
    "41740": ["san-diego-ca", "chula-vista-ca", "carlsbad-ca", "oceanside-ca", "escondido-ca"],
    "38900": ["portland-or", "beaverton-or", "hillsboro-or", "gresham-or"],
    "42660": ["seattle-wa", "tacoma-wa", "bellevue-wa", "kent-wa", "renton-wa", "redmond-wa"],
    "33460": ["minneapolis-mn", "st-paul-mn", "bloomington-mn"],
    "45300": ["tampa-fl", "st-petersburg-fl", "clearwater-fl"],
    "41180": ["st-louis-mo"],
    "12580": ["baltimore-md"],
    "36740": ["orlando-fl"],
    "16740": ["charlotte-nc"],
    "41700": ["san-antonio-tx"],
    "39300": ["providence-ri", "warwick-ri", "cranston-ri"],
    "34980": ["nashville-tn", "murfreesboro-tn", "franklin-tn"],
    "27260": ["jacksonville-fl"],
    "31140": ["louisville-ky"],
    "36420": ["oklahoma-city-ok", "edmond-ok", "norman-ok"],
    "40060": ["richmond-va"],
    "35380": ["new-orleans-la"],
    "39580": ["raleigh-nc", "cary-nc"],
    "32820": ["memphis-tn"],
    "41620": ["salt-lake-city-ut", "west-jordan-ut", "west-valley-city-ut", "sandy-ut"],
    "13820": ["birmingham-al"],
    "26900": ["indianapolis-in", "carmel-in", "fishers-in"],
    "28140": ["kansas-city-mo", "kansas-city-ks", "overland-park-ks", "olathe-ks"],
    "17460": ["cleveland-oh"],
    "18140": ["columbus-oh"],
    "12420": ["austin-tx", "round-rock-tx", "cedar-park-tx"],
    "40900": ["sacramento-ca", "roseville-ca", "elk-grove-ca"],
    "17140": ["cincinnati-oh"],
    "29820": ["las-vegas-nv", "henderson-nv", "north-las-vegas-nv"],
    "38300": ["pittsburgh-pa"],
    "26620": ["des-moines-ia"],
    "33340": ["milwaukee-wi"],
    "36540": ["omaha-ne"],
    "24340": ["grand-rapids-mi"],
    "46060": ["tucson-az"],
    "21340": ["el-paso-tx"],
    "15380": ["buffalo-ny"],
    "40380": ["rochester-ny"],
    "10740": ["albuquerque-nm"],
    "12940": ["baton-rouge-la"],
    "44060": ["spokane-wa"],
    "14260": ["boise-id"],
    "19820": ["detroit-mi", "warren-mi", "dearborn-mi", "livonia-mi"],
    "24860": ["greenville-sc"],
    "16700": ["charleston-sc", "north-charleston-sc"],
    "25540": ["hartford-ct"],
    "35300": ["new-haven-ct"],
    "10900": ["allentown-pa", "bethlehem-pa"],
    "46140": ["tulsa-ok"],
    "47260": ["virginia-beach-va", "norfolk-va", "newport-news-va"],
    "24660": ["greensboro-nc"],
    "27140": ["jackson-ms"],
    "30460": ["lexington-ky"],
    "10420": ["akron-oh"],
    "19380": ["dayton-oh"],
    "16860": ["chattanooga-tn"],
    "25060": ["harrisburg-pa"],
    "30780": ["little-rock-ar"],
    "20500": ["durham-nc"],
    "44700": ["stockton-ca"],
    "42540": ["scranton-pa"],
    "49340": ["worcester-ma"],
    "15980": ["cape-coral-fl"],
    "29460": ["lakeland-fl"],
    "48620": ["wichita-ks"],
    "10580": ["albany-ny"],
    "15540": ["burlington-vt"],
    "11700": ["asheville-nc"],
    "22020": ["fargo-nd"],
    "28940": ["knoxville-tn"],
    "17820": ["colorado-springs-co"],
    "26180": ["honolulu-hi"],
    "11260": ["anchorage-ak"],
    "43580": ["sioux-falls-sd"],
    "41940": ["san-jose-ca", "sunnyvale-ca", "santa-clara-ca", "mountain-view-ca", "palo-alto-ca"],
    "41884": ["san-francisco-ca", "oakland-ca", "berkeley-ca", "fremont-ca", "hayward-ca"],
}

# All SOC codes we care about
RELEVANT_SOCS = set(SLUG_TO_SOC.values())

def main():
    print("═══════════════════════════════════════════════")
    print("  BLS OEWS Data Integration Script (Python)")
    print("═══════════════════════════════════════════════\n")

    script_dir = Path(__file__).parent
    root_dir = script_dir.parent
    xlsx_path = root_dir / "all_data_M_2024.xlsx"

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        print("Download 'All data' from https://www.bls.gov/oes/tables.htm")
        sys.exit(1)

    print("Step 1: Loading Excel file (read-only streaming mode)...")
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    print(f"  ✓ Opened sheet: {ws.title}")

    # Read header row
    headers = []
    header_map = {}  # lowercase -> column index
    national_salaries = {}  # soc -> median
    metro_salaries = {}  # area_code -> soc -> median

    row_count = 0
    national_count = 0
    metro_count = 0

    print("\nStep 2: Scanning rows (this may take a few minutes for large files)...")

    for row in ws.iter_rows(values_only=True):
        if not headers:
            headers = [str(h).strip() if h else "" for h in row]
            header_map = {h.upper(): i for i, h in enumerate(headers)}
            print(f"  Columns: {', '.join(headers[:15])}...")
            continue

        row_count += 1
        if row_count % 500000 == 0:
            print(f"  ... processed {row_count:,} rows")

        # Get column values using header mapping
        def get_val(col_names):
            for name in col_names:
                idx = header_map.get(name.upper())
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
            return ""

        area_type = get_val(["AREA_TYPE", "AREATYPE"])
        area = get_val(["AREA", "AREA_CODE"])
        occ_code = get_val(["OCC_CODE"])
        a_median = get_val(["A_MEDIAN", "ANNUAL_MEDIAN"])

        # Skip if not a SOC code we care about
        if occ_code not in RELEVANT_SOCS:
            continue

        # Parse median salary
        if not a_median or a_median in ("*", "#", "N/A", ""):
            continue
        try:
            salary = int(float(a_median.replace(",", "").replace("$", "")))
        except ValueError:
            continue
        if salary <= 0:
            continue

        # Determine if national or metro
        is_national = (area_type == "1" or area.startswith("99"))

        # Normalize area code (remove leading zeros)
        area_clean = area.lstrip("0") or area

        # Metro: area_type 2 or 4, or 5-digit MSA codes
        is_metro = (area_type in ("2", "4") or
                    (len(area_clean) == 5 and area_clean.isdigit() and not area_clean.startswith("99")))

        if is_national:
            national_salaries[occ_code] = salary
            national_count += 1
        elif is_metro:
            if area_clean not in metro_salaries:
                metro_salaries[area_clean] = {}
            metro_salaries[area_clean][occ_code] = salary
            metro_count += 1

    wb.close()
    print(f"\n  ✓ Processed {row_count:,} total rows")
    print(f"  ✓ National salaries: {len(national_salaries)} occupations ({national_count} relevant rows)")
    print(f"  ✓ Metro salaries: {len(metro_salaries)} areas ({metro_count:,} relevant data points)")

    # Step 3: Load our job data
    print("\nStep 3: Loading Pulsafi job data...")
    job_data_path = root_dir / "app" / "data" / "jobSalaryData.js"
    city_data_path = root_dir / "app" / "data" / "cityData.js"

    # Parse job slugs from jobSalaryData.js
    job_slugs = list(SLUG_TO_SOC.keys())
    print(f"  ✓ {len(job_slugs)} job slugs mapped to SOC codes")

    # Step 4: Build national median updates
    print("\nStep 4: Mapping national BLS data to jobs...")
    job_updates = {}
    unmatched = []

    for slug, soc in SLUG_TO_SOC.items():
        if soc in national_salaries:
            job_updates[slug] = {
                "median": national_salaries[soc],
                "soc": soc
            }
        else:
            unmatched.append(f"{slug} (SOC {soc})")

    print(f"  ✓ Matched {len(job_updates)}/{len(SLUG_TO_SOC)} jobs")
    if unmatched:
        print(f"  ⚠ Unmatched ({len(unmatched)}): {', '.join(unmatched[:10])}")

    # Step 5: Build metro overrides
    print("\nStep 5: Building metro-level overrides...")

    # Load city slugs from cityData.js to validate
    city_data_content = city_data_path.read_text()
    valid_cities = set()
    import re
    for match in re.finditer(r'"([a-z0-9-]+)":\s*\{', city_data_content):
        valid_cities.add(match.group(1))
    print(f"  ✓ Found {len(valid_cities)} cities in cityData.js")

    metro_overrides = {}
    override_count = 0
    cities_with_data = 0

    for area_code, city_slugs in METRO_TO_CITIES.items():
        area_data = metro_salaries.get(area_code, {})
        if not area_data:
            continue

        for city_slug in city_slugs:
            if city_slug not in valid_cities:
                continue

            city_overrides = {}
            for slug, soc in SLUG_TO_SOC.items():
                if soc in area_data:
                    city_overrides[slug] = area_data[soc]
                    override_count += 1

            if city_overrides:
                metro_overrides[city_slug] = city_overrides
                cities_with_data += 1

    print(f"  ✓ {override_count:,} metro-level salary overrides")
    print(f"  ✓ {cities_with_data} cities have BLS metro data")

    # Step 6: Write output
    print("\nStep 6: Writing blsSalaryData.js...")
    output_path = root_dir / "app" / "data" / "blsSalaryData.js"

    from datetime import date
    today = date.today().isoformat()

    output = f"""/**
 * BLS Salary Data — Generated {today}
 *
 * Source: Bureau of Labor Statistics, Occupational Employment and Wage Statistics (OEWS)
 * May 2024 estimates (released April 2025)
 * https://www.bls.gov/oes/
 *
 * National median salaries: Updated from BLS data for {len(job_updates)}/{len(SLUG_TO_SOC)} occupations
 * Metro-level overrides: {override_count:,} salary data points across {cities_with_data} cities
 *
 * DO NOT EDIT MANUALLY — regenerate with: python3 scripts/fetch-bls-data.py
 */

// Metro-level median salary overrides (citySlug -> jobSlug -> annual median salary)
const blsMetroSalaries = {json.dumps(metro_overrides, indent=2)};

// Updated national median salaries from BLS (jobSlug -> {{ median, soc }})
const blsNationalMedians = {json.dumps(job_updates, indent=2)};

module.exports = {{ blsMetroSalaries, blsNationalMedians }};
"""

    output_path.write_text(output)
    print(f"  ✓ Written to app/data/blsSalaryData.js")

    # Summary
    print("\n═══════════════════════════════════════════════")
    print("  Summary")
    print("═══════════════════════════════════════════════")
    print(f"  Jobs with BLS national data:  {len(job_updates)}/{len(SLUG_TO_SOC)}")
    print(f"  Metro-level overrides:        {override_count:,}")
    print(f"  Cities with metro data:       {cities_with_data}")
    print()
    print("  Next steps:")
    print("  1. Run: node scripts/update-page-for-bls.js")
    print("  2. Commit and push")

if __name__ == "__main__":
    main()
